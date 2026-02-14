import os
import uuid
import base64
import io
import mimetypes
import requests
import json
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PILImage
from dotenv import load_dotenv
from supabase import create_client, Client

import database

load_dotenv()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB max

# Supabase Init
url: str = os.environ.get("SUPABASE_URL")
key: str = os.environ.get("SUPABASE_KEY")
supabase: Client = create_client(url, key) if url and key else None


# ── Pages ──────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


# ── Name CRUD API ──────────────────────────────────────────

@app.route('/api/names', methods=['GET'])
def api_get_names():
    names = database.get_all_names()
    return jsonify(names)


@app.route('/api/names', methods=['POST'])
def api_create_name():
    data = request.json
    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'Name is required'}), 400
    
    # Check if exists (simple check, though DB has unique constraint usually)
    existing = supabase.table('names').select("*").eq('name', data['name'].strip()).execute()
    if existing.data:
        return jsonify({'error': 'Name already exists'}), 409

    result = database.create_name(data['name'].strip())
    if result is None:
        return jsonify({'error': 'Failed to create name'}), 500
    return jsonify(result), 201


@app.route('/api/names/<int:name_id>', methods=['PUT'])
def api_update_name(name_id):
    data = request.json
    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'Name is required'}), 400
    result = database.update_name(name_id, data['name'].strip())
    if result is None:
        return jsonify({'error': 'Name not found or failed update'}), 404
    return jsonify(result)


@app.route('/api/names/<int:name_id>', methods=['DELETE'])
def api_delete_name(name_id):
    success = database.delete_name(name_id)
    if not success:
         return jsonify({'error': 'Failed to delete'}), 500
    return jsonify({'success': True})


# ── Entry API ──────────────────────────────────────────────

@app.route('/api/entries', methods=['GET'])
def api_get_entries():
    entries = database.get_all_entries()
    # Sign URLs for photos or use public URLs
    for entry in entries:
        # Assuming bucket is public for simplicity, or we generate signed URL
        # For this app, let's assume public bucket 'photos'
        if entry.get('photo_filename'):
            entry['photo_url'] = f"{url}/storage/v1/object/public/photos/{entry['photo_filename']}"
    return jsonify(entries)


@app.route('/api/entries', methods=['POST'])
def api_create_entry():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    name = data.get('name', '').strip()
    photos = data.get('photos') # Expecting list of base64 strings
    # Backward compatibility
    if not photos and data.get('photo'):
        photos = [data.get('photo')]

    if not name or not photos:
        return jsonify({'error': 'Name and at least one photo are required'}), 400

    uploaded_filenames = []
    photo_urls = []

    try:
        for i, photo_base64 in enumerate(photos):
            # 1. Decode generic base64
            if ',' in photo_base64:
                header, encoded = photo_base64.split(',', 1)
            else:
                encoded = photo_base64

            img_bytes = base64.b64decode(encoded)
            
            # 2. Open with Pillow to compress/resize
            img = PILImage.open(io.BytesIO(img_bytes))
            
            # Convert RGBA to RGB if needed
            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')

            # Resize if too large (max 1600px width)
            max_width = 1600
            if img.width > max_width:
                ratio = max_width / img.width
                new_height = int(img.height * ratio)
                img = img.resize((max_width, new_height), PILImage.Resampling.LANCZOS)
            
            # Compress to JPEG
            out_io = io.BytesIO()
            img.save(out_io, format='JPEG', quality=70, optimize=True)
            out_bytes = out_io.getvalue()

            # Generate filename
            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Use a clean name and unique ID
            filename = f"{name.replace(' ', '_')}_{timestamp_str}_{i}.jpg"

            # Upload to Supabase Storage
            res = supabase.storage.from_("photos").upload(
                filename,
                out_bytes,
                {"content-type": "image/jpeg"}
            )
            # Check for errors omitted as per previous fix

            uploaded_filenames.append(filename)
            photo_urls.append(f"{url}/storage/v1/object/public/photos/{filename}")

    except Exception as e:
        return jsonify({'error': f'Image processing/upload failed: {str(e)}'}), 500

    # Build payload for n8n — n8n handles DB insert, Drive upload, Sheet append
    payload = {
        'name': name,
        'date': data.get('date', datetime.now().strftime('%Y-%m-%d')),
        'timestamp': data.get('timestamp', datetime.now().strftime('%H:%M:%S')),
        'latitude': data.get('latitude', 0),
        'longitude': data.get('longitude', 0),
        'area_name': data.get('area_name', ''),
        'extracted_text': data.get('extracted_text', ''),
        'photo_urls': photo_urls,
        'photo_filenames': uploaded_filenames
    }

    # Send to n8n webhook (synchronous — wait for n8n to finish)
    webhook_url = os.environ.get('N8N_WEBHOOK_URL')
    if not webhook_url:
        return jsonify({'error': 'N8N_WEBHOOK_URL not configured'}), 500

    try:
        n8n_response = requests.post(webhook_url, json=payload, timeout=30)
        
        if n8n_response.status_code >= 400:
            # Log full details server-side
            print(f"n8n webhook error [{n8n_response.status_code}]: {n8n_response.text}")
            
            # Try to extract a meaningful error message from n8n's response
            try:
                n8n_err = n8n_response.json()
                err_msg = n8n_err.get('message', n8n_err.get('error', n8n_response.text[:200]))
            except Exception:
                err_msg = n8n_response.text[:200] if n8n_response.text else f'HTTP {n8n_response.status_code}'
            
            return jsonify({
                'error': f'Workflow failed: {err_msg}',
                'status_code': n8n_response.status_code
            }), 500
        
        # Return success with the payload data
        result = payload.copy()
        result['main_photo_url'] = photo_urls[0] if photo_urls else None
        return jsonify(result), 201

    except requests.exceptions.Timeout:
        # n8n took too long, but photos are already in Supabase Storage
        return jsonify({'warning': 'Entry submitted but workflow timed out', **payload}), 202
    except requests.exceptions.ConnectionError:
        print(f"Cannot reach n8n server at {webhook_url}")
        return jsonify({'error': 'Cannot reach n8n server — is it running?'}), 503
    except Exception as e:
        print(f"Failed to call n8n webhook: {e}")
        return jsonify({'error': f'Webhook call failed: {str(e)}'}), 500


@app.route('/api/entries/<int:entry_id>', methods=['DELETE'])
def api_delete_entry(entry_id):
    row = database.delete_entry(entry_id)
    if row:
        try:
            supabase.storage.from_("photos").remove([row['photo_filename']])
        except Exception as e:
            print(f"Failed to delete file {row['photo_filename']}: {e}")
            
    return jsonify({'success': True})


# ── Excel Export ───────────────────────────────────────────

@app.route('/api/export', methods=['GET'])
def api_export():
    try:
        dates = database.get_all_dates()
    except Exception:
        return jsonify({'error': 'Database connection failed'}), 500
        
    if not dates:
        return jsonify({'error': 'No entries to export'}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = "Site Follow-up"

    # Styles
    date_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    date_font = Font(bold=True, size=14)
    header_fonts = Font(bold=True, size=11)
    time_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    name_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    loc_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    photo_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30

    current_row = 1

    for date_str in dates:
        entries = database.get_entries_by_date(date_str)
        if not entries:
            continue

        # Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        date_cell = ws.cell(row=current_row, column=1, value=f"Date : {date_str}")
        date_cell.fill = date_fill
        date_cell.font = date_font
        date_cell.alignment = center_align
        ws.cell(row=current_row, column=2).fill = date_fill
        current_row += 1

        headers = [('Time', time_fill), ('Name', name_fill), ('Location', loc_fill), ('Photo', photo_fill)]
        for col_idx, (header, fill) in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.fill = fill
            cell.font = header_fonts
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

        for entry in entries:
            ws.row_dimensions[current_row].height = 80
            
            # Text Cells
            ws.cell(row=current_row, column=1, value=entry['timestamp']).alignment = center_align
            ws.cell(row=current_row, column=1).border = thin_border
            
            ws.cell(row=current_row, column=2, value=entry['name']).alignment = center_align
            ws.cell(row=current_row, column=2).border = thin_border
            
            loc_text = entry['area_name'] if entry.get('area_name') else f"{entry.get('location_lat')}, {entry.get('location_lng')}"
            ws.cell(row=current_row, column=3, value=loc_text).alignment = center_align
            ws.cell(row=current_row, column=3).border = thin_border
            
            ws.cell(row=current_row, column=4).border = thin_border

        # Download and embed photo
        if entry.get('photo_filename'):
            try:
                # Handle both old (string) and new (JSON list) formats
                filenames_str = entry['photo_filename']
                filename_to_use = None
                
                try:
                    filenames = json.loads(filenames_str)
                    if isinstance(filenames, list) and len(filenames) > 0:
                        filename_to_use = filenames[0] # Use first photo for Excel
                except json.JSONDecodeError:
                    filename_to_use = filenames_str # Old format

                if filename_to_use:
                    # Download bytes from Supabase Storage
                    res = supabase.storage.from_("photos").download(filename_to_use)
                    img_bytes = res # return value is bytes
                    
                    img = XlImage(io.BytesIO(img_bytes))
                    img.width = 100
                    img.height = 75
                    ws.add_image(img, f"D{current_row}")
            except Exception as e:
                print(f"Error embedding image {entry['photo_filename']}: {e}")
                ws.cell(row=current_row, column=4, value="[Image Load Failed]")

            current_row += 1
        
        current_row += 1

    # Save to BytesIO for memory download
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Site_Followup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ── Vercel Handler ─────────────────────────────────────────
# No special handler needed for Flask on Vercel if using wsgi

if __name__ == '__main__':
    print("🚀 Starting Server on http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)
